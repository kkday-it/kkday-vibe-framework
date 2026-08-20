"""
Insurance App - Web UI
使用 Streamlit 取代原本的 Tkinter，讓 DevOps 可以在 EKS 部署網頁版。
"""
import streamlit as st
import datetime
import os
import openpyxl
import tempfile
# 在這個框架中，我們會透過 platform_sdk 呼叫 workflow
from platform_sdk import run_workflow

st.set_page_config(page_title="🏖️ 旅行業責任險自動投保", page_icon="🏖️")
st.title("🏖️ 旅行業責任險自動投保系統")

st.markdown("### 📅 旅遊期間設定")
col1, col2 = st.columns(2)
with col1:
    start_date = st.date_input("起保日期", datetime.date.today() + datetime.timedelta(days=1))
with col2:
    end_date = st.date_input("迄保日期", datetime.date.today() + datetime.timedelta(days=2))

st.markdown("### 📁 名單上傳")
uploaded_file = st.file_uploader("上傳保險名單 Excel", type=["xlsx"])

if st.button("🚀 開始投保作業", type="primary"):
    if not uploaded_file:
        st.error("請先上傳名單 Excel！")
    elif end_date <= start_date:
        st.error("迄保日期必須大於起保日期！")
    else:
        # 將日期轉成民國年
        eff_date = f"{start_date.year - 1911}/{start_date.month:02d}/{start_date.day:02d}"
        exp_date = f"{end_date.year - 1911}/{end_date.month:02d}/{end_date.day:02d}"
        
        # 1. 將上傳的檔案存入暫存目錄 (取代原本寫死的 EXCEL_DIR)
        temp_dir = tempfile.mkdtemp()
        excel_path = os.path.join(temp_dir, uploaded_file.name)
        with open(excel_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        
        # 2. 解析 Excel 資訊
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            last_name = str(ws['D2'].value or "").strip()
            first_name = str(ws['E2'].value or "").strip()
            represent_mem = f"{last_name} {first_name}".strip()
            total_mem = 0
            for row in ws.iter_rows(min_col=1, max_col=1):
                for cell in row:
                    if cell.value is not None:
                        try:
                            val = int(cell.value)
                            if val > 0:
                                total_mem = val
                        except:
                            pass
            tra_no = os.path.splitext(uploaded_file.name)[0]
            
            st.info(f"解析成功：代表人 {represent_mem}，共 {total_mem} 人，圈號 {tra_no}")
        except Exception as e:
            st.error(f"Excel 解析失敗：{str(e)}")
            st.stop()

        # 3. 呼叫底層自動化 Workflow
        with st.spinner("自動化執行中，這可能需要幾分鐘的時間，請稍候..."):
            result = run_workflow(
                workflow_id="submit_insurance",
                inputs={
                    "eff_date": eff_date,
                    "exp_date": exp_date,
                    "excel_file_path": excel_path,
                    "tra_no": tra_no,
                    "represent_mem": represent_mem,
                    "total_mem": total_mem
                }
            )
            
            if result.get("status") == "success":
                st.success("✅ 投保完成！")
                st.markdown(f"""
                ### 📄 投保結果
                - **交易序號**：{result.get('transaction_no')}
                - **覆核號碼**：{result.get('review_no')}
                - **保險費**：{result.get('insurance_fee')} 元
                - **記錄檔**：`{result.get('txt_path')}`
                - **保單 PDF**：`{result.get('pdf_path', '下載失敗，請手動確認')}`
                """)
            else:
                st.error(f"❌ 投保失敗：{result.get('error_msg')}")
